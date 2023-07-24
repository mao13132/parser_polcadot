from openpyxl import Workbook
from openpyxl.styles import Font


class SaveResult:
    def __init__(self, good_dict, ivent_dict):
        self.good_dict = good_dict

        self.ivent_dict = ivent_dict

        self.colums = ['name_post', 'link', 'date', 'name_them', 'post_text', 'name_author', 'likes_post',
                       'dislike_post', 'status', 'model', 'decision', 'confirmation', 'Votes in favor', 'Votes against',
                       'DOTs in favor', 'DOTs against', 'support', 'issuance', 'author_comment', 'text_comment', 'like_comment',
                       'dislike_comment', 'time_comment']

        self.colums_ivent = ['date', 'time', 'author', 'name']

        self.comment_colums = 19

    def create_title(self, ws):
        for count, col in enumerate(self.colums):

            ws.cell(row=1, column=count + 1).value = col
            ws.cell(row=1, column=count + 1).font = Font(bold=True)

    def merge_row(self, ws, count_comments, count_def):

        for count in range(self.comment_colums - 1):
            count = count + 1
            ws.merge_cells(start_column=count, start_row=count_def,
                           end_column=count, end_row=count_def + count_comments - 1)

        return True

    def write_data(self, ws, count_def, post, name_them):


        ws.cell(row=count_def, column=1).value = post['name_post']
        ws.cell(row=count_def, column=2).value = post['link']
        ws.cell(row=count_def, column=3).value = post['date']
        ws.cell(row=count_def, column=4).value = name_them
        ws.cell(row=count_def, column=5).value = post['data']['post_text']
        ws.cell(row=count_def, column=6).value = post['data']['name_author']
        ws.cell(row=count_def, column=7).value = int(post['data']['like_post'])
        ws.cell(row=count_def, column=8).value = int(post['data']['dislike_post'])
        ws.cell(row=count_def, column=9).value = post['data']['status']
        ws.cell(row=count_def, column=10).value = post['data']['model']
        ws.cell(row=count_def, column=11).value = post['data']['decision']
        ws.cell(row=count_def, column=12).value = post['data']['confirmation']
        ws.cell(row=count_def, column=13).value = post['data']['aye']
        ws.cell(row=count_def, column=14).value = post['data']['nay']
        ws.cell(row=count_def, column=15).value = post['data']['ayes']
        ws.cell(row=count_def, column=16).value = post['data']['nays']
        ws.cell(row=count_def, column=17).value = post['data']['support']
        ws.cell(row=count_def, column=18).value = post['data']['issuance']

        if len(post['data']['comment']) == 0:
            return True

        for count_com, comment in enumerate(post['data']['comment']):
            ws.cell(row=count_def + count_com, column=19).value = comment['name_comment']
            ws.cell(row=count_def + count_com, column=20).value = comment['text_comment']
            ws.cell(row=count_def + count_com, column=21).value = int(comment['like'])
            ws.cell(row=count_def + count_com, column=22).value = int(comment['dislike'])
            ws.cell(row=count_def + count_com, column=23).value = comment['time_comment']





    def itter_rows(self, ws):
        count_def = 2  # с двойки т.к. в 1 строке уже заголовок
        for count_them, them in enumerate(self.good_dict):
            for count_post, post in enumerate(them['links']):
                count_comments = len(post['data']['comment'])


                if count_comments > 1:

                    response = self.merge_row(ws, count_comments, count_def)


                write_data = self.write_data(ws, count_def, post, them['name_them'])


                if count_comments > 1:
                    count_def = count_def + count_comments
                else:
                    count_def += 1


    def one_sheet(self, ws):

        response = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def write_title_ivents(self, ws):
        for count, col in enumerate(self.colums_ivent):

            ws.cell(row=1, column=count + 1).value = col
            ws.cell(row=1, column=count + 1).font = Font(bold=True)

        return True

    def write_ivent(self, count_def, ws, ivent):

        ws.cell(row=count_def, column=1).value = ivent['data_com']
        ws.cell(row=count_def, column=2).value = ivent['time_comm']
        ws.cell(row=count_def, column=3).value = ivent['author_comm']
        ws.cell(row=count_def, column=4).value = ivent['text_comm']

        return True


    def itter_ivents(self, ws):
        count_def = 2

        for count_ivent, ivent in enumerate(self.ivent_dict):

            self.write_ivent(count_def, ws, ivent)

            count_def += 1


    def save_ivent(self, ws):

        self.write_title_ivents(ws)

        self.itter_ivents(ws)



    def save_file(self, filename):

        wb = Workbook()

        ws = wb.active

        result = self.one_sheet(ws)

        wb.create_sheet('ivent')

        ws = wb['ivent']

        result_ivent = self.save_ivent(ws)

        filename = f'{filename}.xlsx'

        wb.save(filename)

        return filename


if __name__ == '__main__':
    from src.temp_list import TempList

    result = SaveResult(TempList.good_ower, TempList.temp_ivent).save_file('1111')

    print(result)
